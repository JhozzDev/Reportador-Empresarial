import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook


root = tk.Tk()
root.withdraw()

carpeta = filedialog.askdirectory(title="Selecciona la carpeta con los reportes")

if not carpeta:
    print("No se seleccionó ninguna carpeta")
    exit()


archivos = [
    os.path.join(carpeta, f)
    for f in os.listdir(carpeta)
    if f.endswith(".xlsx") and not f.startswith("~$")
]

if not archivos:
    print("No se encontraron archivos Excel")
    exit()

print(f"{len(archivos)} archivos encontrados")

reporte_mensual = None
FILA_INICIO = 13
FILA_FIN = 52
COL_INICIO = 1
COL_FIN = 13

for archivo in archivos:
    df = pd.read_excel(archivo)


    #Convierte los valores a numero
    df = df.apply(pd.to_numeric, errors="coerce")

    df_rango = df.iloc[FILA_INICIO:FILA_FIN, COL_INICIO:COL_FIN]

    if reporte_mensual is None:
        reporte_mensual = df_rango
    else:
        reporte_mensual = reporte_mensual.add(df_rango)


wb = load_workbook("plantilla.xlsx")
ws = wb.active


for i, fila in enumerate(reporte_mensual.values, start=15):
    for j, valor in enumerate(fila, start=2):
        ws.cell(row=i, column=j, value=valor)

wb.save("Reporte_mensual.xlsx")

print("Reporte mensual creado correctamente")